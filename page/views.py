from django.shortcuts import render, get_object_or_404
from item.models import *
from openpyxl import load_workbook
def index(request):
    all_categories = Category.objects.filter(is_active=True, is_in_index_catalog=True)
    return render(request, 'page/index.html', locals())


def about(request):
    return render(request, 'page/about.html', locals())

def cats(request):
    wb = load_workbook(filename='C:/Users/ххх/PycharmProjects/ideaHome/cats.xlsx')
    sheet = wb.active

    max_row = sheet.max_row

    max_column = sheet.max_column
    for i in range(1, max_row + 1):
        # worksheet.write('A{}'.format(row), cat_id)
        # worksheet.write('B{}'.format(row), cat_name)
        # worksheet.write('C{}'.format(row), cat_parent_id)
        # worksheet.write('D{}'.format(row), cat_description)
        # worksheet.write('E{}'.format(row), cat_img)
        # worksheet.write('F{}'.format(row), cat_title)
        # worksheet.write('G{}'.format(row), cat_keywords)
        # worksheet.write('H{}'.format(row), cat_meta_description)
        old_id=sheet.cell(row=i, column=1).value
        cat_name=sheet.cell(row=i, column=2).value
        cat_parent_id=sheet.cell(row=i, column=3).value
        try:
            cat_description=sheet.cell(row=i, column=4).value.replace('_x000D_','')
        except:
            cat_description = sheet.cell(row=i, column=4).value
        cat_img=sheet.cell(row=i, column=5).value
        cat_title=sheet.cell(row=i, column=6).value
        cat_keywords=sheet.cell(row=i, column=7).value
        cat_meta_description=sheet.cell(row=i, column=8).value
        # if cat_parent_id ==0:
        #     Category.objects.create(old_id=old_id,name=cat_name,description=cat_description,image='images/catalog/categories/'+cat_img,
        #                             page_title=cat_title,page_description=cat_meta_description,page_keywords=cat_keywords)
        if cat_parent_id != 0:
            cat = Category.objects.get(old_id=cat_parent_id)
            SubCategory.objects.create(old_id=old_id, name=cat_name, description=cat_description, category=cat,
                                    page_title=cat_title, page_description=cat_meta_description,
                                    page_keywords=cat_keywords)
    return render(request, 'page/about.html', locals())



def manuf(request):
    wb = load_workbook(filename='C:/Users/ххх/PycharmProjects/ideaHome/manuf.xlsx')
    sheet = wb.active

    max_row = sheet.max_row

    max_column = sheet.max_column
    for i in range(1, max_row + 1):
        # worksheet.write('A{}'.format(row), cat_id)
        # worksheet.write('B{}'.format(row), cat_name)
        # worksheet.write('C{}'.format(row), cat_img)
        # worksheet.write('D{}'.format(row), cat_title)
        # worksheet.write('E{}'.format(row), cat_keywords)
        # worksheet.write('F{}'.format(row), cat_meta_description)
        old_id=sheet.cell(row=i, column=1).value
        name=sheet.cell(row=i, column=2).value
        img=sheet.cell(row=i, column=3).value
        cat_title=sheet.cell(row=i, column=4).value
        cat_keywords=sheet.cell(row=i, column=5).value
        cat_meta_description=sheet.cell(row=i, column=6).value

        Manufactor.objects.create(old_id=old_id, name=name, image='images/catalog/manufacturers/'+img,
                                    page_title=cat_title, page_description=cat_meta_description,
                                    page_keywords=cat_keywords)
    return render(request, 'page/about.html', locals())

def itemm(request):
    wb = load_workbook(filename='C:/Users/ххх/PycharmProjects/ideaHome/items.xlsx')
    sheet = wb.active

    max_row = sheet.max_row

    max_column = sheet.max_column
    for i in range(1, max_row + 1):
        # worksheet.write('A{}'.format(row), item_id)
        # worksheet.write('B{}'.format(row), item_cat_id)
        # worksheet.write('C{}'.format(row), item_name)
        # worksheet.write('D{}'.format(row), item_articul)
        # worksheet.write('E{}'.format(row), item_description)
        # worksheet.write('F{}'.format(row), item_manufactor)
        # worksheet.write('G{}'.format(row), item_unit)
        # worksheet.write('H{}'.format(row), item_price)
        # worksheet.write('I{}'.format(row), item_img_main)
        # worksheet.write('J{}'.format(row), item_img_add)
        # worksheet.write('K{}'.format(row), item_title)
        # worksheet.write('L{}'.format(row), item_keywords)
        # worksheet.write('M{}'.format(row), item_meta_description)
        old_id=sheet.cell(row=i, column=1).value
        item_cat_id=sheet.cell(row=i, column=2).value
        item_name=sheet.cell(row=i, column=3).value
        item_articul=sheet.cell(row=i, column=4).value
        try:
            item_description = sheet.cell(row=i, column=5).value.replace('_x000D_', '')
        except:
            item_description = sheet.cell(row=i, column=5).value
        item_manufactor = sheet.cell(row=i, column=6).value
        item_unit = sheet.cell(row=i, column=7).value
        item_price = sheet.cell(row=i, column=8).value
        item_img_main = sheet.cell(row=i, column=9).value
        item_img_add = sheet.cell(row=i, column=10).value
        item_title = sheet.cell(row=i, column=11).value
        item_keywords = sheet.cell(row=i, column=12).value
        item_meta_description = sheet.cell(row=i, column=13).value
        cat= None
        subcat = None


        try:
            cat = Category.objects.get(old_id=item_cat_id)

        except:
            subcat = SubCategory.objects.get(old_id=item_cat_id)


        try:
            manufactor = Manufactor.objects.get(old_id=item_manufactor)
        except:
            manufactor=None
        print(manufactor)
        print(old_id)
        if item_img_main:
            item_first_big_img= item_img_main.split('|')[0]
            item_first_small_img = item_img_main.split('|')[2]
        else:
            item_first_big_img = None
            item_first_small_img = None
        item = None
        if cat:
            item = Item.objects.create(manufactor=manufactor,category=cat, name=item_name,price=item_price,article=item_articul,
                                       units=item_unit,old_id=old_id,description=item_description,
                                       page_title=item_title,page_description=item_meta_description,page_keywords=item_keywords)
        if subcat:
            catt = subcat.category
            item = Item.objects.create(manufactor=manufactor,category=catt, subcategory=subcat, name=item_name, price=item_price, article=item_articul,
                                       units=item_unit, old_id=old_id, description=item_description,
                                       page_title=item_title, page_description=item_meta_description,
                                       page_keywords=item_keywords)
        if item_img_main:
            ItemImage.objects.create(item=item,image='images/catalog/items/'+item_first_big_img,image_small='/media/images/catalog/items/'+item_first_small_img)
        if item_img_add:
            for img in item_img_add.splitlines():
                item_big_img = item_img_main.split('|')[0]
                item_small_img = item_img_main.split('|')[1]
                ItemImage.objects.create(item=item, image='images/catalog/items/'+item_big_img, image_small='/media/images/catalog/items/'+item_small_img)


    return render(request, 'page/about.html', locals())

def contacts(request):
    return render(request, 'page/contacts.html', locals())


def delivery(request):
    return render(request, 'page/delivery.html', locals())


def manufacturers(request):
    all_manufacturers = Manufactor.objects.all()
    return render(request, 'page/manufacturers.html', locals())

def catalog(request):
    all_categories = Category.objects.all()
    return render(request, 'page/catalog.html', locals())


def manufacturers_cat(request,slug):
    return render(request, 'page/manufacturers.html', locals())



def category(request, category_slug):
    all_categories = Category.objects.filter(is_active=True)
    category = Category.objects.get(name_slug=category_slug)
    items_qs = Item.objects.filter(category=category)
    items = items_qs
    manufactors = category.get_all_manufactors()
    qs_filtered = False
    search_res = False

    order = request.GET.get('order')
    filter_manufactor = request.GET.get('manufactor')
    search_q = request.GET.get('q')


    if search_q:
        items = items_qs.filter(name_lower__contains=search_q.lower())
        search_res = True
        param_search = search_q

    if filter_manufactor:
        if search_res:
            items = items.filter(manufactor__name_slug=filter_manufactor)
        else:
            items = items_qs.filter(manufactor__name_slug=filter_manufactor)
        qs_filtered = True
        param_manufactor = filter_manufactor

    if order == 'price_gte':
        if qs_filtered or search_res:
            items = items.order_by('-price')
        else:
            items = items_qs.order_by('-price')
        param_order = order

    if order == 'price_lte':
        if qs_filtered or search_res:
            items = items.order_by('price')
        else:
            items = items_qs.order_by('price')
        param_order = order

    if order == 'name_az':
        if qs_filtered or search_res:
            items = items.order_by('name')
        else:
            items = items_qs.order_by('name')
        param_order = order

    if order == 'name_za':
        if qs_filtered or search_res:
            items = items.order_by('-name')
        else:
            items = items_qs.order_by('-name')
        param_order = order
    return render(request, 'page/category.html', locals())

def subcategory(request, category_slug, subcategory_slug):
    all_categories = Category.objects.filter(is_active=True)
    category = Category.objects.get(name_slug=category_slug)
    subcategory = SubCategory.objects.get(name_slug=subcategory_slug)
    items_qs = Item.objects.filter(subcategory=subcategory)
    items = items_qs
   # manufactors = subcategory.manufactor_set.all()
    qs_filtered = False
    search_res = False

    order = request.GET.get('order')
    filter_manufactor = request.GET.get('manufactor')
    search_q = request.GET.get('q')

    if search_q:
        items = items_qs.filter(name_lower__contains=search_q.lower())
        search_res = True
        param_search = search_q

    if filter_manufactor:
        if search_res:
            items = items.filter(manufactor__name_slug=filter_manufactor)
        else:
            items = items_qs.filter(manufactor__name_slug=filter_manufactor)
        qs_filtered = True
        param_manufactor = filter_manufactor

    if order == 'price_gte':
        if qs_filtered or search_res:
            items = items.order_by('-price')
        else:
            items = items_qs.order_by('-price')
        param_order = order

    if order == 'price_lte':
        if qs_filtered or search_res:
            items = items.order_by('price')
        else:
            items = items_qs.order_by('price')
        param_order = order

    if order == 'name_az':
        if qs_filtered or search_res:
            items = items.order_by('name')
        else:
            items = items_qs.order_by('name')
        param_order = order

    if order == 'name_za':
        if qs_filtered or search_res:
            items = items.order_by('-name')
        else:
            items = items_qs.order_by('-name')
        param_order = order

    return render(request, 'page/category.html', locals())

def item_page(request,category_slug,item_slug,subcategory_slug=None):
    if subcategory_slug:
        subcategory = get_object_or_404(SubCategory, name_slug=subcategory_slug)
    category = get_object_or_404(Category, name_slug=category_slug)


    item = get_object_or_404(Item, name_slug=item_slug)

    filter_qs(1,None,{'dsf':'sada'})
    return render(request, 'page/item.html', locals())

def filter_qs(qs,*args,**kwargs):
    print('filter')
    print(qs)
    print(**kwargs)